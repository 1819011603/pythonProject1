import openpyxl
import pymysql
import os


def export_to_excel_xlsx(worksheet, cursor, table):
    # 首先向excel表中写入数据表的字段  desc 表名；
    column_count = cursor.execute("desc %s" % table)  # 列数
    for i in range(1, column_count + 1):
        temp_tuple = cursor.fetchone()  # 单行
        # print(temp_tuple)
        # print(temp_tuple)  #列属性
        # Field        | Type         | Null | Key | Default | Extra
        # id           | int unsigned | NO   | PRI | NULL    | auto_increment
        # ('id', 'int', 'NO', 'PRI', None, '')
        # 行列索引都从1开始从而不是0
        worksheet.cell(1, i).value = temp_tuple[0]  # （0，i，temp_tuple[0]） 第1行i列 值为列名

    row_count = cursor.execute("select * from %s" % table)  # 行数
    for i in range(1, row_count + 1):
        temp_tuple = cursor.fetchone()  # 拿出一行
        # print(temp_tuple)
        for j in range(1, column_count + 1):  # 遍历每行的每列

            worksheet.cell(i + 1, j).value = temp_tuple[j - 1]  # (i + 1, j, temp_tuple[j])  i+1行  j列  写入i+1行j列的数据
            # 第一行是列名


# 数据库写入到excel中
def database_to_excel_xlsx(database, directory_name="database", sheet="sheet"):
    """

    :param database: 数据库名称
    :param directory_name: 文件夹名称
    :param sheet: excel的工作sheet名称
    """
    # 连接数据库
    global workbook
    connect = pymysql.Connect(
        host='127.0.0.1',  # 主机名
        port=3306,  # 端口号
        user='root',  # 用户名
        passwd='abc123',  # 密码
        charset="utf8mb4",  # 默认编码
        db=database  # 数据库名称
    )

    # 创建一个游标
    cursor = connect.cursor()
    cursor.execute("show tables")
    tables = cursor.fetchall()  # tables is tuple  fetchall()取出操作返回的所有的行
    table_list = [tuple[0] for tuple in tables]
    directory_name += "/" + database  # 文件夹路径

    if os.path.exists(directory_name) is False and os.path.isdir(directory_name) is False:
        os.makedirs(directory_name)
    path = os.getcwd() + "/" + directory_name + "/"
    for table_name in table_list:
        print("database_name: " + database + ", table_name: " + table_name + ",\tpath: " + path + table_name + ".xlsx")
        # 创建一个workbook 设置编码
        workbook = openpyxl.Workbook()
        # 当前的sheet
        worksheet = workbook.active
        # 创建一个worksheet
        # worksheet = workbook.create_sheet()

        # 写入worksheet
        export_to_excel_xlsx(worksheet, cursor, table_name)
        workbook.save(path + table_name + ".xlsx")

    cursor.close()
    workbook.close()
    connect.close()


if __name__ == "__main__":
    database_name = "water"  # 数据库名称
    database_to_excel_xlsx(database_name)
