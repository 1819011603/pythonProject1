import matplotlib.pyplot as plt
import pandas
import xlrd
import openpyxl
import numpy

# xlrd只能支持xls
# excel = pandas.read_excel("intensity_info.xls")
# print(excel)


t = openpyxl.load_workbook("intensity_info.xlsx", True).active
ans = []
for i in t.values:
    ans.append(i[3])

ans = ans[1:]
temp = []
for s in ans:
    temp.append([j for j in s.split(",")])

temp = numpy.array(temp, dtype=numpy.int32)

p = openpyxl.load_workbook("spectra_info.xlsx", True).active

str = p.cell(2, 4).value.split(",")
str = numpy.array(str, dtype=numpy.float)
print(str)
u = openpyxl.load_workbook("sample_info.xlsx", True).active
temp1 = []
for s in u["E2":"E44"]:
    temp1.append(s[0].value)
for i in range(9, 10):
    plt.plot(str, temp[i])
plt.xlabel(u"g")
plt.ylabel(u"G")
plt.show()
