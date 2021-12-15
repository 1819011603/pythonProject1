import pandas as pd
import openpyxl

intensity = openpyxl.load_workbook("intensity_info.xlsx", True).active
column = ["id", "sample_id"]

spectra = openpyxl.load_workbook("spectra_info.xlsx", True).active

spectra_weight = spectra.cell(2, 4).value.split(",")
for i in spectra_weight:
    column.append(float(i))

column.append("humidity")
print(column)
workbook = openpyxl.Workbook()
worksheet = workbook.active

for c in range(len(column)):
    worksheet.cell(1,c+1).value = column[c]

sample = openpyxl.load_workbook("sample_info.xlsx", True).active
s = []
for i in sample['E1':'E44']:
    s.append(i[0].value)

print(s)
u = [i for i in intensity.values]
temp = []
t = 0
row = 2
for i in u[1:]:
    temp.append(i[0])
    temp.append(i[1])
    for j in i[3].split(","):
        temp.append(int(j))
    temp.append(s[int(t / 10) + 1])
    t += 1

    for u in range(len(temp)):
        worksheet.cell(row,u+1).value = temp[u]
    row += 1
    print(temp)
    temp = []
workbook.save("test_all.xlsx")

