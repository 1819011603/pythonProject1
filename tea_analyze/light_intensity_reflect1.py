import pandas as pd
import openpyxl

intensity = openpyxl.load_workbook("intensity_info.xlsx", True).active
column = ["sample_id"]

spectra = openpyxl.load_workbook("spectra_info.xlsx", True).active

spectra_weight = spectra.cell(2, 4).value.split(",")
for i in spectra_weight:
    column.append(float(i))
stand_light = openpyxl.load_workbook("standlight_info.xlsx", True).active

spectra_light = stand_light.cell(3, 4).value.split(",")
light = []
for i in spectra_light:
    light.append(float(i))
spectra_dark = spectra.cell(2, 6).value.split(",")
dark = []
for i in spectra_dark:
    dark.append(float(i))
column.append("humidity")

workbook = openpyxl.Workbook()
worksheet = workbook.active

for c in range(len(column)):
    worksheet.cell(1,c+1).value = column[c]

sample = openpyxl.load_workbook("sample_info.xlsx", True).active
s = []
for i in sample['E2':'E44']:
    s.append(i[0].value)


u = [i for i in intensity.values]
temp = []

row = 2
for i in u[1:]:

    temp.append(i[1])
    for j in i[3].split(","):
        temp.append(float(j))
    for i in range(1,1+len(temp[1:])):
        temp[i] = round((round(temp[i],1)-dark[i-1])/(round(light[i-1],1) - dark[i-1]),4)
    temp.append(float(s[(row-2)//10]))
    print(temp)
    for u in range(len(temp)):
        worksheet.cell(row, u + 1).value = temp[u]
    row += 1
    temp=[]
workbook.save("test_all_reflect1.xlsx")

